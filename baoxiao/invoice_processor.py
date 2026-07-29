"""发票 PDF 分类、字段解析、行程单配对和 Excel/报销单输出。"""

from pathlib import Path
import re
import shutil

import fitz
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

import itinerary
import ordinary_invoice
import train_ticket


CATEGORIES = ("高铁票", "出租票", "住宿票", "其他发票")
OUTPUT_FOLDERS = CATEGORIES + ("行程单", "非发票")


def extract_text(pdf_path):
    document = fitz.open(pdf_path)
    try:
        return "\n".join(page.get_text() for page in document)
    finally:
        document.close()


def classify(filename, text):
    """先识别票面类型，再在普通发票内部分类，禁止把任意 PDF 当发票。"""
    if train_ticket.is_train_ticket_text(text):
        return "高铁票"
    if itinerary.is_itinerary_text(text):
        return "行程单"
    if ordinary_invoice.is_ordinary_invoice_text(text):
        return ordinary_invoice.classify_invoice(f"{filename}\n{text}")
    return "非发票"


def _empty_row(path, category):
    return {
        "category": category, "filename": path.name, "date": "", "time": "",
        "route": "", "amount": "", "invoice_no": "未识别", "kind": "未识别",
        "hotel": "未识别",
    }


def parse_pdf(pdf_path):
    path = Path(pdf_path)
    text = extract_text(path)
    category = classify(path.name, text)
    if category == "高铁票":
        row = _empty_row(path, category)
        ticket = train_ticket.extract_ticket(str(path))
        row.update(
            date=ticket["date"], time=ticket["time"], route=ticket["route"],
            amount=ticket["refundFee"] if float(ticket["refundFee"]) else ticket["amount"],
            invoice_no=ticket["invoiceNo"], kind="铁路电子客票",
        )
        return row
    if category == "行程单":
        return itinerary.parse_itinerary(path)
    if category in ("出租票", "住宿票", "其他发票"):
        row = _empty_row(path, category)
        parsed = ordinary_invoice.parse_invoice(text)
        row.update(parsed)
        if category == "出租票":
            row["invoice_date"] = row["date"]
        return row
    row = _empty_row(path, "非发票")
    row["warning"] = "未识别为铁路电子客票、普通电子发票或行程单"
    return row


def _unique_path(directory, filename):
    destination = Path(directory) / filename
    counter = 1
    while destination.exists():
        destination = Path(directory) / f"{Path(filename).stem}_{counter}{Path(filename).suffix}"
        counter += 1
    return destination


def _tokens(filename):
    stem = Path(filename).stem.lower()
    for word in ("电子发票", "发票", "行程单", "行程", "明细", "invoice", "trip"):
        stem = stem.replace(word, " ")
    return {token for token in re.split(r"[^0-9a-z\u4e00-\u9fff]+", stem) if len(token) >= 2}


def pair_taxi_itineraries(taxis, itineraries):
    """一张出租发票至多匹配一张行程单；低置信度时保持未匹配。"""
    remaining = list(itineraries)
    for taxi in taxis:
        candidates = []
        for trip_sheet in remaining:
            shared_tokens = _tokens(taxi["filename"]) & _tokens(trip_sheet["filename"])
            score = len(shared_tokens) * 3
            if taxi.get("amount") and taxi["amount"] == trip_sheet.get("amount"):
                score += 10
            if taxi.get("invoice_date") and taxi["invoice_date"] == trip_sheet.get("application_date"):
                score += 2
            candidates.append((score, trip_sheet))
        best = max(candidates, key=lambda item: item[0]) if candidates else None
        if best and best[0] >= 5:
            trip_sheet = best[1]
            taxi["date"] = trip_sheet.get("date", "")
            taxi["route"] = trip_sheet.get("route", "")
            taxi["trips"] = trip_sheet.get("trips", [])
            taxi["itinerary_filename"] = trip_sheet["filename"]
            remaining.remove(trip_sheet)
        else:
            # 开票日期不是实际乘车日期；没有行程单时不写入报销日期和路线。
            taxi["date"] = ""
            taxi["route"] = ""
            taxi["trips"] = []
            taxi["itinerary_filename"] = "未匹配"
    return remaining


def categorize_pdfs(pdf_files, output_dir):
    root = Path(output_dir)
    for folder in OUTPUT_FOLDERS:
        (root / folder).mkdir(parents=True, exist_ok=True)
    rows = []
    for pdf_path in pdf_files:
        path = Path(pdf_path)
        if path.suffix.lower() != ".pdf" or not path.exists():
            continue
        try:
            row = parse_pdf(path)
        except Exception as exc:
            row = _empty_row(path, "非发票")
            row["warning"] = f"PDF 解析失败: {exc}"
        destination = _unique_path(root / row["category"], path.name)
        shutil.move(str(path), destination)
        row["filename"], row["path"] = destination.name, str(destination)
        rows.append(row)
        print(f"  {path.name} -> {row['category']}")

    taxis = [row for row in rows if row["category"] == "出租票"]
    itineraries = [row for row in rows if row["category"] == "行程单"]
    unmatched = pair_taxi_itineraries(taxis, itineraries)
    for taxi in taxis:
        if taxi["itinerary_filename"] == "未匹配":
            print(f"  警告: 出租票 {taxi['filename']} 未匹配行程单")
    if unmatched:
        print(f"  警告: {len(unmatched)} 份行程单未匹配出租票")

    results = {folder: [] for folder in OUTPUT_FOLDERS}
    for row in rows:
        results[row["category"]].append(row["filename"])
    return results, [row for row in rows if row["category"] in CATEGORIES]


HEADER_FILL = PatternFill("solid", fgColor="4472C4")
TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
SEPARATOR_FILL = PatternFill("solid", fgColor="D9EAF7")


def _style_headers(sheet, row_number, headers):
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(row_number, column, header)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_sheet(workbook, title, headers, rows, fields):
    sheet = workbook.create_sheet(title)
    sheet.freeze_panes = "A2"
    _style_headers(sheet, 1, headers)
    amount_column = fields.index("amount") + 1
    for row_number, row in enumerate(rows, 2):
        for column, field in enumerate(fields, 1):
            value = row.get(field, "")
            if field == "amount" and value:
                value = float(value)
            cell = sheet.cell(row_number, column, value)
            if field == "invoice_no" and value and value != "未识别":
                cell.value = "\u200b" + str(value)
            if field in ("invoice_no", "filename"):
                cell.number_format = "@"
            if field in ("date", "route", "hotel", "kind", "filename"):
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.cell(row_number, amount_column).number_format = "0.00"
        line_count = max(
            str(row.get("date", "")).count("、") + 1,
            str(row.get("route", "")).count("；") + 1,
        )
        sheet.row_dimensions[row_number].height = max(32, line_count * 20)
    if rows:
        sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(headers)).column_letter}{len(rows) + 1}"
    total_row = len(rows) + 2
    sheet.cell(total_row, amount_column - 1, "合计")
    amount_cell = sheet.cell(total_row, amount_column)
    amount_cell.value = f"=SUM({get_column_letter(amount_column)}2:{get_column_letter(amount_column)}{len(rows) + 1})" if rows else 0
    amount_cell.number_format = "0.00"
    for cell in sheet[total_row][:len(headers)]:
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True, color="375623")
    _finish_sheet(sheet)
    widths = {
        "出租票": (24, 58, 12, 24, 58),
        "住宿票": (14, 38, 12, 24, 58),
        "其他发票": (14, 30, 12, 24, 48),
    }
    for index, width in enumerate(widths[title], 1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_train_sheet(workbook, rows):
    sheet = workbook.create_sheet("高铁票")
    sheet.freeze_panes = "A2"
    headers = ["日期", "行程", "金额", "发票号", "文件名"]
    _style_headers(sheet, 1, headers)
    output_row = 1
    previous_date = None
    for row in rows:
        date = row.get("date", "")
        if previous_date is not None and date != previous_date:
            output_row += 1
            for cell in sheet[output_row][:5]:
                cell.fill = SEPARATOR_FILL
            sheet.row_dimensions[output_row].height = 10
        output_row += 1
        values = (date, row.get("route", ""), float(row["amount"]) if row.get("amount") else "", row.get("invoice_no", "未识别"), row["filename"])
        for column, value in enumerate(values, 1):
            cell = sheet.cell(output_row, column, value)
            if column == 4 and value and value != "未识别":
                cell.value = "\u200b" + str(value)
            if column in (4, 5):
                cell.number_format = "@"
            if column in (2, 5):
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.cell(output_row, 3).number_format = "0.00"
        sheet.row_dimensions[output_row].height = 30
        previous_date = date

    total_row = output_row + 1
    sheet.cell(total_row, 2, "合计")
    sheet.cell(total_row, 3, f"=SUM(C2:C{output_row})" if rows else 0)
    sheet.cell(total_row, 3).number_format = "0.00"
    for cell in sheet[total_row][:5]:
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True, color="375623")
    if rows:
        sheet.auto_filter.ref = f"A1:E{output_row}"
    _finish_sheet(sheet)
    for index, width in enumerate((14, 32, 12, 24, 42), 1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _finish_sheet(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            alignment = cell.alignment
            cell.alignment = Alignment(
                horizontal=alignment.horizontal,
                vertical="center",
                text_rotation=alignment.text_rotation,
                wrap_text=alignment.wrap_text,
                shrink_to_fit=alignment.shrink_to_fit,
                indent=alignment.indent,
            )
    for column_index, column in enumerate(sheet.columns, 1):
        letter = get_column_letter(column_index)
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[letter].width = min(max(width, 12), 48)
    sheet.sheet_view.showGridLines = True


def _write_combined_sheet(workbook, rows):
    sheet = workbook.create_sheet("合并统计")
    sheet.merge_cells("A1:F1")
    title = sheet["A1"]
    title.value = "全部发票费用统计"
    title.font = Font(size=16, bold=True, color="FFFFFF")
    title.fill = HEADER_FILL
    title.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30

    _style_headers(sheet, 3, ["类别", "张数", "金额"])
    detail_start = 11
    detail_end = max(detail_start, detail_start + len(rows) - 1)
    for row_number, category in enumerate(CATEGORIES, 4):
        sheet.cell(row_number, 1, category)
        sheet.cell(row_number, 2, f'=COUNTIF($A${detail_start}:$A${detail_end},A{row_number})')
        sheet.cell(row_number, 3, f'=SUMIF($A${detail_start}:$A${detail_end},A{row_number},$D${detail_start}:$D${detail_end})')
        sheet.cell(row_number, 3).number_format = "0.00"
    sheet.cell(8, 1, "总计")
    sheet.cell(8, 2, "=SUM(B4:B7)")
    sheet.cell(8, 3, "=SUM(C4:C7)")
    sheet.cell(8, 3).number_format = "0.00"
    for cell in sheet[8][:3]:
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True, color="375623")

    headers = ["类别", "日期", "内容", "金额", "发票号", "文件名"]
    _style_headers(sheet, 10, headers)
    for row_number, row in enumerate(rows, detail_start):
        if row["category"] == "住宿票":
            content = row.get("hotel", "")
        elif row["category"] == "其他发票":
            content = row.get("kind", "")
        else:
            content = row.get("route") or row.get("kind", "")
        values = (row["category"], row.get("date", ""), content, float(row["amount"]) if row.get("amount") else "", row.get("invoice_no", "未识别"), row["filename"])
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_number, column, value)
            if column == 5 and value != "未识别":
                cell.value = "\u200b" + str(value)
                cell.number_format = "@"
            if column in (2, 3, 6):
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.cell(row_number, 4).number_format = "0.00"
        line_count = max(str(values[1]).count("、") + 1, str(values[2]).count("；") + 1)
        sheet.row_dimensions[row_number].height = max(32, line_count * 20)
    if rows:
        sheet.auto_filter.ref = f"A10:F{detail_end}"
    sheet.freeze_panes = "A11"
    for index, width in enumerate((14, 24, 58, 14, 25, 58), 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            alignment = cell.alignment
            cell.alignment = Alignment(
                horizontal=alignment.horizontal,
                vertical="center",
                text_rotation=alignment.text_rotation,
                wrap_text=alignment.wrap_text,
                shrink_to_fit=alignment.shrink_to_fit,
                indent=alignment.indent,
            )
    sheet.sheet_view.showGridLines = True


def generate_excel(rows, excel_path):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    grouped = {category: [] for category in CATEGORIES}
    for row in rows:
        grouped[row["category"]].append(row)
    for category_rows in grouped.values():
        category_rows.sort(key=lambda row: (row.get("date", ""), row.get("time", ""), row["filename"]))
    combined_rows = sorted(rows, key=lambda row: (row.get("date", ""), row["category"], row["filename"]))
    _write_combined_sheet(workbook, combined_rows)
    _write_train_sheet(workbook, grouped["高铁票"])
    _write_sheet(workbook, "出租票", ["日期", "行程", "金额", "发票号", "文件名"], grouped["出租票"], ["date", "route", "amount", "invoice_no", "filename"])
    _write_sheet(workbook, "住宿票", ["日期", "酒店", "金额", "发票号", "文件名"], grouped["住宿票"], ["date", "hotel", "amount", "invoice_no", "filename"])
    _write_sheet(workbook, "其他发票", ["日期", "种类", "金额", "发票号", "文件名"], grouped["其他发票"], ["date", "kind", "amount", "invoice_no", "filename"])
    workbook.save(excel_path)

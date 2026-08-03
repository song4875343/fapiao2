"""High-speed-rail based taxi trip planning and fixed-template exports."""

from pathlib import Path

from openpyxl import load_workbook


def _city(value):
    value = str(value or "").strip()
    return value.replace("站", "")


def _train_legs(trains, home_city, home_label="家", station_label="高铁站", project_label="项目所在地"):
    legs = []
    seen_journeys = set()
    for train in trains:
        if train.get("is_refund") or train.get("isRefund"):
            continue
        route = str(train.get("route") or "")
        parts = [part.strip() for part in route.split(" - ", 1)]
        if len(parts) != 2:
            continue
        start, end = parts
        unit = _city(home_city)
        start_city, end_city = _city(start), _city(end)
        if start_city == unit or start_city.startswith(unit):
            direction = "outbound"
            candidates = [(home_label, station_label, 1), (station_label, project_label, 2)]
        elif end_city == unit or end_city.startswith(unit):
            direction = "return"
            candidates = [(project_label, station_label, 4), (station_label, home_label, 5)]
        else:
            continue
        journey_key = (train.get("date", ""), direction)
        if journey_key in seen_journeys:
            continue
        seen_journeys.add(journey_key)
        for index, (origin, destination, phase) in enumerate(candidates):
            legs.append({
                "target_date": train.get("date", ""),
                "origin": origin,
                "destination": destination,
                "train_route": route,
                "leg_index": index + 1,
                "phase": phase,
                "direction": direction,
            })
    return sorted(legs, key=lambda item: (item["target_date"], item["direction"], item["leg_index"]))


def _flatten_taxis(taxis):
    records = []
    for taxi in taxis:
        trips = taxi.get("trips") or []
        if not trips:
            trips = [{
                "date": taxi.get("invoice_date", ""), "time": "",
                "start": taxi.get("route", "").split(" - ", 1)[0],
                "end": taxi.get("route", "").split(" - ", 1)[-1],
                "amount": taxi.get("amount", ""),
            }]
        for index, trip in enumerate(trips):
            records.append({
                "taxi": taxi, "trip": trip, "trip_index": index,
                "has_original_trip": bool(taxi.get("trips")),
                "invoice_date": taxi.get("invoice_date", "") or trip.get("date", ""),
                "amount": float(trip.get("amount") or 0),
                "original_date": trip.get("date", "") or taxi.get("invoice_date", ""),
                "original_start": trip.get("start", ""), "original_end": trip.get("end", ""),
            })
    return records


def plan_taxis(taxis, trains, home_city, home_to_station=(20.0, 30.0), station_to_project_min=30.0,
               home_label="家", station_label="高铁站", project_label="项目所在地"):
    """Return planned rows and mutate taxi rows so summary reflects planned routes."""
    records = _flatten_taxis(taxis)
    usable_trains = [train for train in trains if not (train.get("is_refund") or train.get("isRefund"))]
    legs = _train_legs(usable_trains, home_city, home_label, station_label, project_label)
    used = set()
    planned = []

    def assign(record, leg):
        status = "改写" if record["has_original_trip"] else "抄写"
        record["trip"]["date"] = leg["target_date"]
        record["trip"]["start"] = leg["origin"]
        record["trip"]["end"] = leg["destination"]
        record["trip"]["status"] = status
        record["trip"]["note"] = status
        if not record["taxi"].get("trips"):
            record["taxi"]["date"] = leg["target_date"]
            record["taxi"]["route"] = f"{leg['origin']} - {leg['destination']}"
            record["taxi"]["status"] = status
            record["taxi"]["note"] = status
        planned.append({
            **leg, "time": record["trip"].get("time", ""), "status": status,
            "note": status, "amount": record["amount"], "source": record,
        })

    # Base station-transfer legs are matched only by their configured amount
    # rules. An unmatched base leg is omitted from the final form.
    for leg in legs:
        is_home_leg = home_label in (leg["origin"], leg["destination"])
        lo, hi = home_to_station if is_home_leg else (station_to_project_min, float("inf"))
        choices = []
        for idx, record in enumerate(records):
            eligible = lo <= record["amount"] <= hi if is_home_leg else record["amount"] > lo
            if idx in used or not eligible:
                continue
            target_amount = (lo + hi) / 2 if hi != float("inf") else lo
            choices.append((abs(record["amount"] - target_amount), record["amount"], idx))
        if not choices:
            continue
        _, _, idx = min(choices)
        used.add(idx)
        assign(records[idx], leg)

    # Consume every remaining taxi amount with unrestricted project-to-project
    # legs, distributed across train tickets in chronological round-robin order.
    train_dates = sorted({train.get("date", "") for train in usable_trains if train.get("date")})
    remaining = [(idx, record) for idx, record in enumerate(records) if idx not in used]
    for extra_index, (idx, record) in enumerate(remaining):
        target_date = train_dates[extra_index % len(train_dates)] if train_dates else record["original_date"]
        leg = {
            "target_date": target_date,
            "origin": "项目地点1",
            "destination": "项目地点2",
            "train_route": "",
            "leg_index": 3 + extra_index,
            "phase": 3,
            "direction": "extra",
        }
        used.add(idx)
        assign(record, leg)
    return planned


def _fill_form(sheet, rows, start_row, end_row, title_row, date_row, total_row, note_row,
               total_rows=None, write_total=True, sequence_start=1):
    for row in range(start_row, end_row + 1):
        for col in range(1, 9):
            sheet.cell(row, col).value = None
    seen_sources = set()
    for offset, item in enumerate(rows[: end_row - start_row + 1]):
        row = start_row + offset
        source = item.get("source")
        source_key = id(source) if source is not None else None
        ticket_count = 1 if source_key is None or source_key not in seen_sources else ""
        if source_key is not None:
            seen_sources.add(source_key)
        values = (sequence_start + offset, 1, item.get("target_date", ""), item.get("origin", ""),
                  item.get("destination", ""), item.get("amount", ""), ticket_count,
                  item.get("note", ""))
        for col, value in enumerate(values, 1):
            sheet.cell(row, col).value = value
    total = sum(float(item.get("amount") or 0) for item in (total_rows if total_rows is not None else rows))
    if write_total:
        sheet.cell(total_row, 6).value = total
        sheet.cell(note_row, 1).value = f"合计人民币（大写）：{total:.2f}"
    else:
        sheet.cell(total_row, 6).value = None
        sheet.cell(note_row, 1).value = None
    sheet.cell(date_row, 7).value = f"票据 {sum(1 for item in rows if item.get('amount'))} 张"


def export_fixed_templates(rows, single_template, double_template, single_path=None, double_path=None):
    """Write the immutable double-page template; single_path is retained for API compatibility."""
    double_template = Path(double_template)
    rows = sorted(rows, key=lambda item: (
        item.get("target_date", ""), item.get("phase", 3),
        item.get("leg_index", 0), item.get("time", ""),
        item.get("origin", ""), item.get("destination", ""),
    ))
    wb = load_workbook(double_template)
    chunks = [rows[index:index + 24] for index in range(0, len(rows), 24)] or [[]]
    sheets = [wb.active]
    for _ in chunks[1:]:
        sheets.append(wb.copy_worksheet(wb.active))
    for index, (sheet, chunk) in enumerate(zip(sheets, chunks), 1):
        sheet.title = f"出租车报销单{index}"
        sequence_start = (index - 1) * 24 + 1
        if len(chunk) <= 12:
            _fill_form(sheet, chunk, 6, 17, 1, 3, 18, 19, sequence_start=sequence_start)
        else:
            _fill_form(sheet, chunk[:12], 6, 17, 1, 3, 18, 19, write_total=False,
                       sequence_start=sequence_start)
            _fill_form(sheet, chunk[12:], 27, 38, 22, 24, 39, 40, total_rows=chunk,
                       sequence_start=sequence_start + 12)
    if double_path:
        wb.save(double_path)

"""使用现有 XLSX 模板生成出租车报销单。"""

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill


ROWS_PER_SHEET = 12
DATA_START_ROW = 6
TOTAL_ROW = DATA_START_ROW + ROWS_PER_SHEET
OUTPUT_COLUMNS = 9
SPLIT_COLORS = ("FCE4D6", "E2F0D9", "DDEBF7", "E4DFEC", "FFF2CC")
CN_DIGITS = "零壹贰叁肆伍陆柒捌玖"
CN_UNITS = ("", "拾", "佰", "仟")
CN_BIG_UNITS = ("", "万", "亿", "兆")


def _integer_to_chinese(value):
    if value == 0:
        return "零"
    groups = []
    while value:
        groups.append(value % 10000)
        value //= 10000
    result = ""
    pending_zero = False
    for group_index in range(len(groups) - 1, -1, -1):
        group = groups[group_index]
        if group == 0:
            pending_zero = bool(result)
            continue
        if result and (pending_zero or group < 1000):
            result += "零"
        pending_zero = False
        group_text = ""
        zero_between = False
        for unit_index in range(3, -1, -1):
            digit = group // (10 ** unit_index) % 10
            if digit:
                if zero_between and group_text:
                    group_text += "零"
                group_text += CN_DIGITS[digit] + CN_UNITS[unit_index]
                zero_between = False
            elif group_text:
                zero_between = True
        result += group_text + CN_BIG_UNITS[group_index]
    return result


def money_to_chinese(amount):
    cents = int(round(float(amount) * 100))
    integer, fraction = divmod(cents, 100)
    result = _integer_to_chinese(integer) + "元"
    jiao, fen = divmod(fraction, 10)
    if not fraction:
        return result + "整"
    if jiao:
        result += CN_DIGITS[jiao] + "角"
    elif fen:
        result += "零"
    if fen:
        result += CN_DIGITS[fen] + "分"
    return result


def _split_route(route):
    parts = route.split(" - ", 1)
    return (parts[0], parts[1]) if len(parts) == 2 else (route, "")


def _expense_rows(taxis):
    rows = []
    split_group_index = 0
    for taxi in taxis:
        trips = taxi.get("trips") or []
        if trips:
            ordered_trips = sorted(
                trips,
                key=lambda trip: (
                    not bool(trip.get("date")), trip.get("date", ""), trip.get("time", "")
                ),
            )
            is_split = len(ordered_trips) > 1
            split_source = float(taxi["amount"]) if is_split and taxi.get("amount") else ""
            split_color = SPLIT_COLORS[split_group_index % len(SPLIT_COLORS)] if is_split else ""
            if is_split:
                split_group_index += 1
            for index, trip in enumerate(ordered_trips):
                rows.append({
                    "date": trip["date"], "start": trip["start"], "end": trip["end"],
                    "time": trip.get("time", ""), "amount": float(trip["amount"]),
                    "ticket_count": 1 if index == 0 else "", "split_source": split_source,
                    "purpose": trip.get("note") or trip.get("status") or "",
                    "split_color": split_color,
                })
        else:
            start, end = _split_route(taxi.get("route", ""))
            rows.append({
                "date": taxi.get("date", ""), "start": start, "end": end,
                "time": taxi.get("time", ""),
                "amount": float(taxi["amount"]) if taxi.get("amount") else "", "ticket_count": 1,
                "split_source": "", "split_color": "",
                "purpose": taxi.get("note") or taxi.get("status") or "",
            })
    rows.sort(key=lambda item: (
        not bool(item["date"]), item["date"], item["time"], item["start"], item["end"]
    ))
    return rows


def _add_split_source_column(sheet):
    for merged_range in ("A1:H1", "A2:H2", "A4:H4", "A19:H19"):
        sheet.unmerge_cells(merged_range)
    purpose_width = sheet.column_dimensions["H"].width
    sheet.insert_cols(8)
    for row_number in range(1, sheet.max_row + 1):
        sheet.cell(row_number, 8)._style = copy(sheet.cell(row_number, 9)._style)
    sheet.column_dimensions["H"].width = 13
    sheet.column_dimensions["I"].width = purpose_width
    sheet.merge_cells("A1:I1")
    sheet.merge_cells("A2:I2")
    sheet.merge_cells("A4:I4")
    sheet.merge_cells("A19:I19")
    sheet["H5"] = "拆分源票额"


def _expand_data_area(sheet, row_count):
    extra_rows = max(0, row_count - ROWS_PER_SHEET)
    if not extra_rows:
        return TOTAL_ROW

    for merged_range in ("A18:B18", "A19:I19", "A20:B20"):
        sheet.unmerge_cells(merged_range)
    sheet.insert_rows(TOTAL_ROW, amount=extra_rows)

    source_row = TOTAL_ROW - 1
    for row_number in range(TOTAL_ROW, TOTAL_ROW + extra_rows):
        sheet.row_dimensions[row_number].height = sheet.row_dimensions[source_row].height
        for column in range(1, OUTPUT_COLUMNS + 1):
            sheet.cell(row_number, column)._style = copy(sheet.cell(source_row, column)._style)

    total_row = TOTAL_ROW + extra_rows
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    sheet.merge_cells(start_row=total_row + 1, start_column=1, end_row=total_row + 1, end_column=OUTPUT_COLUMNS)
    sheet.merge_cells(start_row=total_row + 2, start_column=1, end_row=total_row + 2, end_column=2)
    return total_row


def _fill_sheet(sheet, rows, invoice_dates):
    sheet.title = "出租车报销单"
    total_row = _expand_data_area(sheet, len(rows))
    for row_number in range(DATA_START_ROW, total_row):
        for column in range(1, OUTPUT_COLUMNS + 1):
            sheet.cell(row_number, column).value = None
    for offset, item in enumerate(rows):
        row_number = DATA_START_ROW + offset
        values = (
            offset + 1, 1, item["date"], item["start"], item["end"], item["amount"],
            item["ticket_count"], item["split_source"], item.get("purpose", ""),
        )
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_number, column, value)
            cell.alignment = copy(cell.alignment)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.cell(row_number, 6).number_format = "0.00"
        sheet.cell(row_number, 8).number_format = "0.00"
        if item["split_color"]:
            row_fill = PatternFill("solid", fgColor=item["split_color"])
            for column in range(1, OUTPUT_COLUMNS + 1):
                sheet.cell(row_number, column).fill = row_fill
        address_lines = max(
            (len(str(item["start"])) + 12) // 13,
            (len(str(item["end"])) + 12) // 13,
        )
        sheet.row_dimensions[row_number].height = max(20, address_lines * 20)

    total = sum(float(item["amount"]) for item in rows if item["amount"] != "")
    ticket_count = sum(1 for item in rows if item["ticket_count"] == 1)
    sheet.cell(total_row, 6, total)
    sheet.cell(total_row, 6).number_format = "0.00"
    sheet["G3"] = f"票据 {ticket_count} 张"
    sheet.cell(total_row + 1, 1, f"合计人民币（大写）：{money_to_chinese(total)}    ￥{total:.2f}")
    sheet.print_area = f"A1:I{total_row + 2}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    dates = sorted(date for date in invoice_dates if date)
    if dates:
        year, month, day = dates[-1].split("-")
        sheet["D3"] = f"{year} 年 {int(month)} 月 {int(day)} 日"


def generate_taxi_reimbursement(taxis, template_path, output_path):
    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f"出租车报销单模板不存在: {template}")
    workbook = load_workbook(template)
    sheet = workbook.active
    _add_split_source_column(sheet)
    rows = _expense_rows(taxis)
    invoice_dates = [taxi.get("invoice_date", "") for taxi in taxis]
    _fill_sheet(sheet, rows, invoice_dates)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output

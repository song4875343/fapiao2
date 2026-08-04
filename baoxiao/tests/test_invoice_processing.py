import base64
import shutil
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
try:
    from PyPDF2 import PdfReader, PdfWriter
except ImportError:
    from pypdf import PdfReader, PdfWriter

import invoice_processor
import ordinary_invoice
import taxi_reimbursement
import taxi_planner
import reimbursement_workbook
import pdfm_v3


ROOT = Path(__file__).resolve().parents[1]
EXAMPLES = ROOT / "exm"


class InvoiceProcessingTests(unittest.TestCase):
    def test_planner_consumes_every_taxi_with_unrestricted_extra_legs(self):
        trains = [{"date": "2026-07-14", "route": "郑州东 - 上海虹桥"}]
        amounts = (25, 45, 8, 12, 90)
        taxis = [
            {"filename": f"taxi-{index}.pdf", "invoice_date": "2026-07-01",
             "amount": str(amount), "trips": []}
            for index, amount in enumerate(amounts)
        ]
        planned = taxi_planner.plan_taxis(taxis, trains, "郑州")
        self.assertEqual(len(planned), len(taxis))
        self.assertEqual(sum(row["amount"] for row in planned), sum(amounts))
        self.assertFalse(any(row["status"] == "未处理" for row in planned))
        extras = [row for row in planned if row["direction"] == "extra"]
        self.assertTrue(all(row["origin"] == "项目地点1" for row in extras))
        self.assertTrue(all(row["destination"] == "项目地点2" for row in extras))

    def test_planner_ignores_refunds_and_deduplicates_same_day_direction(self):
        trains = [
            {"date": "2026-05-20", "route": "北京西 - 郑州东", "is_refund": True},
            {"date": "2026-05-20", "route": "北京西 - 郑州东", "is_refund": False},
            {"date": "2026-05-20", "route": "北京西 - 郑州东", "is_refund": False},
        ]
        taxis = [
            {"filename": f"taxi-{index}.pdf", "invoice_date": "2026-05-01",
             "amount": str(amount), "trips": []}
            for index, amount in enumerate((65, 24, 61, 23))
        ]
        planned = taxi_planner.plan_taxis(taxis, trains, "郑州")
        self.assertEqual(sum(row["phase"] == 4 for row in planned), 1)
        self.assertEqual(sum(row["phase"] == 5 for row in planned), 1)
        self.assertEqual(sum(row["direction"] == "extra" for row in planned), 2)

    def test_amount_accepts_currency_symbol_after_value(self):
        text = """价税合计（大写）\n（小写）\n79.99\n¥\n2.33\n¥\n77.66\n¥"""
        self.assertEqual(ordinary_invoice.extract_amount(text), "79.99")

    def test_all_example_classifications_and_fields(self):
        expected = {
            "26119110010006876913.pdf": ("高铁票", "33.00", "26119110010006876913"),
            "26119121152006130026.pdf": ("高铁票", "371.00", "26119121152006130026"),
            "8b98a33e9bb84f8f8292da80ec4c3b54.pdf": ("其他发票", "360.97", "26137000000336474725"),
            "dzfp_26112000003099003616_中磐科工（河南）有限公司_20260727102832.pdf": ("住宿票", "500.00", "26112000003099003616"),
            "【哈哈出行-23.17元-1个行程】高德打车电子发票.pdf": ("出租票", "23.17", "26142000001115972821"),
            "【江南出行-33.42元-2个行程】高德打车电子发票.pdf": ("出租票", "33.42", "26362000001322695201"),
            "【江南出行-33.42元-2个行程】高德打车电子行程单.pdf": ("行程单", "33.42", "未识别"),
            "【火箭出行-23.81元-1个行程】高德打车电子发票.pdf": ("出租票", "23.81", "26117000001133211484"),
            "【风韵出行特选-64.70元-1个行程】高德打车电子发票.pdf": ("出租票", "64.70", "26427000000571999826"),
            "天津西瓜旅游有限责任公司_发票金额469.00元.pdf": ("住宿票", "469.00", "26127000000357075542"),
            "报价函.pdf": ("非发票", "", "未识别"),
        }
        for name, values in expected.items():
            with self.subTest(name=name):
                row = invoice_processor.parse_pdf(EXAMPLES / name)
                self.assertEqual((row["category"], row["amount"], row["invoice_no"]), values)
        other = invoice_processor.parse_pdf(EXAMPLES / "8b98a33e9bb84f8f8292da80ec4c3b54.pdf")
        self.assertEqual(other["kind"], "水果、塑料制品、茶")

    def test_itinerary_extracts_each_trip(self):
        path = EXAMPLES / "【江南出行-33.42元-2个行程】高德打车电子行程单.pdf"
        row = invoice_processor.parse_pdf(path)
        self.assertEqual(len(row["trips"]), 2)
        self.assertEqual(row["trips"][0]["date"], "2025-12-31")
        self.assertEqual(row["trips"][0]["start"], "升龙城2号院(西门)西侧")
        self.assertEqual(row["trips"][0]["end"], "机械工业第六设计研究院有限公司")
        self.assertEqual(row["trips"][1]["amount"], "9.60")

    def test_end_to_end_outputs(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            raw = temp / "raw"
            output = temp / "output"
            shutil.copytree(EXAMPLES, raw)
            results, rows = invoice_processor.categorize_pdfs(list(raw.glob("*.pdf")), output)
            self.assertEqual(len(results["非发票"]), 1)
            self.assertEqual(len(results["高铁票"]), 2)
            taxis = [row for row in rows if row["category"] == "出租票"]
            matched = next(row for row in taxis if "江南出行" in row["filename"])
            self.assertEqual(matched["date"], "2025-12-31、2026-05-27")
            self.assertEqual(len(matched["trips"]), 2)
            self.assertEqual(sum(row["itinerary_filename"] != "未匹配" for row in taxis), 1)

            summary_path = output / "发票.xlsx"
            invoice_processor.generate_excel(rows, summary_path)
            summary = load_workbook(summary_path, data_only=False)
            self.assertEqual(summary.sheetnames, ["合并统计", "高铁票", "出租票", "住宿票", "其他发票"])
            combined = summary["合并统计"]
            self.assertEqual(combined["A1"].value, "全部发票费用统计")
            self.assertEqual(combined["B4"].value, '=COUNTIF($A$11:$A$19,A4)')
            self.assertEqual(combined["C8"].value, "=SUM(C4:C7)")
            self.assertEqual(combined.max_row, 19)
            self.assertAlmostEqual(sum(combined.cell(row, 4).value for row in range(11, 20)), 1879.07, places=2)
            self.assertEqual(combined["C11"].value, "交通运输服务")
            self.assertEqual(combined["C19"].value, "水果、塑料制品、茶")
            train_sheet = summary["高铁票"]
            self.assertEqual(train_sheet["A1"].value, "日期")
            self.assertEqual(train_sheet["A3"].value, None)
            self.assertEqual(train_sheet["A3"].fill.fgColor.rgb, "00D9EAF7")
            self.assertEqual(train_sheet["B5"].value, "合计")
            self.assertEqual(train_sheet["C5"].value, "=SUM(C2:C4)")
            self.assertEqual(train_sheet["D2"].value.lstrip("\u200b"), "26119110010006876913")
            self.assertEqual(summary["住宿票"]["C2"].value, 500)
            self.assertEqual(summary["出租票"]["B6"].value, "合计")
            self.assertEqual(summary["住宿票"]["B4"].value, "合计")
            self.assertEqual(summary["其他发票"]["B3"].value, "合计")
            for worksheet in summary.worksheets:
                self.assertTrue(worksheet.sheet_view.showGridLines)
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            self.assertEqual(cell.alignment.vertical, "center")
            taxi_sheet = summary["出租票"]
            multi_trip_row = next(
                row for row in range(2, taxi_sheet.max_row + 1)
                if "江南出行" in taxi_sheet.cell(row, 5).value
            )
            self.assertGreater(taxi_sheet.row_dimensions[multi_trip_row].height, 32)

            form_path = output / "出租车报销单.xlsx"
            taxi_reimbursement.generate_taxi_reimbursement(taxis, ROOT / "出租车报销单模板.xlsx", form_path)
            form = load_workbook(form_path, data_only=False)
            sheet = form.active
            self.assertEqual(sheet["H5"].value, "拆分源票额")
            self.assertEqual(sheet["I5"].value, "出差事由")
            self.assertEqual(sheet["C6"].value, "2025-12-31")
            self.assertEqual(sheet["D6"].value, "升龙城2号院(西门)西侧")
            self.assertEqual(sheet["E6"].value, "机械工业第六设计研究院有限公司")
            self.assertGreaterEqual(sheet.row_dimensions[6].height, 40)
            self.assertAlmostEqual(sheet["F18"].value, 145.10, places=2)
            self.assertIn("壹佰肆拾伍元壹角", sheet["A19"].value)
            data_rows = list(sheet.iter_rows(min_row=6, max_row=10, values_only=True))
            self.assertEqual(data_rows[0][7], 33.42)
            self.assertEqual(data_rows[1][7], 33.42)
            self.assertTrue(any(row[7] is None for row in data_rows[2:]))

    def test_taxi_reimbursement_expands_multi_trip_invoices_in_one_sheet(self):
        trips = [
            {
                "date": f"2026-07-{index:02d}",
                "start": f"起点{index}",
                "end": f"终点{index}",
                "amount": "20.00",
            }
            for index in range(13, 0, -1)
        ]
        taxis = [{
            "filename": "多行程发票.pdf",
            "invoice_date": "2026-07-28",
            "amount": "260.00",
            "date": "、".join(trip["date"] for trip in trips),
            "trips": trips,
        }]
        with tempfile.TemporaryDirectory() as temp_dir:
            form_path = Path(temp_dir) / "出租车报销单.xlsx"
            taxi_reimbursement.generate_taxi_reimbursement(
                taxis, ROOT / "出租车报销单模板.xlsx", form_path
            )
            form = load_workbook(form_path, data_only=False)
            self.assertEqual(len(form.worksheets), 1)
            written = [
                row
                for sheet in form.worksheets
                for row in sheet.iter_rows(min_row=6, max_row=18, values_only=True)
                if row[0] is not None
            ]
            self.assertEqual(len(written), 13)
            self.assertEqual(sum(row[5] for row in written), 260.00)
            self.assertEqual(sum(row[6] == 1 for row in written), 1)
            self.assertEqual([row[2] for row in written], sorted(row[2] for row in written))
            self.assertTrue(all(row[7] == 260.00 for row in written))
            split_colors = {
                form.active.cell(row_number, 1).fill.fgColor.rgb
                for row_number in range(6, 19)
            }
            self.assertEqual(len(split_colors), 1)
            self.assertEqual(form.active["F19"].value, 260.00)
            self.assertEqual(form.active["G3"].value, "票据 1 张")

    def test_double_page_form_is_editable_and_prints_one_sheet_per_page(self):
        rows = [
            {
                "target_date": f"2026-07-{index:02d}",
                "origin": "起点",
                "destination": "终点",
                "amount": 10,
                "note": "改写",
                "source": {"index": index},
            }
            for index in range(1, 14)
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            form_path = Path(temp_dir) / reimbursement_workbook.FORM_NAME
            taxi_planner.export_fixed_templates(
                rows,
                ROOT / "出租车报销单模板.xlsx",
                ROOT / "出租车报销单双页模板.xlsx",
                double_path=form_path,
            )
            form = load_workbook(form_path, data_only=False)
            self.assertIsNone(form.active["H6"].value)
            self.assertIsNone(form.active["H27"].value)
            self.assertEqual(form.active["A2"].alignment.horizontal, "center")
            self.assertEqual(form.active["A23"].alignment.horizontal, "center")

            payload = reimbursement_workbook.workbook_payload(form_path)
            self.assertEqual(len(payload["sheets"]), 1)
            result = reimbursement_workbook.save_edits(form_path, [
                {"sheet": "出租车报销单1", "cell": "D6", "value": "新起点"},
                {"sheet": "出租车报销单1", "cell": "F6", "value": "99.50"},
            ])
            self.assertTrue(result["success"])
            edited = load_workbook(form_path, data_only=False).active
            self.assertEqual(edited["D6"].value, "新起点")
            self.assertEqual(edited["F6"].value, 99.5)
            self.assertEqual(edited["F39"].value, 219.5)
            self.assertEqual(edited["G3"].value, "票据 12 张")
            self.assertEqual(edited["G24"].value, "票据 1 张")

            printed = reimbursement_workbook.render_pdf(form_path)
            pdf_page = PdfReader(printed["path"]).pages[0]
            self.assertEqual(len(PdfReader(printed["path"]).pages), 1)
            self.assertAlmostEqual(float(pdf_page.mediabox.width), 595.2756, places=2)
            self.assertAlmostEqual(float(pdf_page.mediabox.height), 841.8898, places=2)

    def test_precomposed_reimbursement_page_bypasses_two_up_scaling(self):
        import fitz
        import pdfm_v3

        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "source.pdf"
            output_path = Path(temp_dir) / "output.pdf"
            source = fitz.open()
            page = source.new_page(width=595, height=842)
            page.insert_text((80, 200), "FULL PAGE", fontsize=20)
            source.save(source_path)
            source.close()

            api = pdfm_v3.PDFMergerAPI(mode="local")
            api._merge_invoice_by_pages([{
                "path": str(source_path), "page_index": 0,
                "rotation": 0, "full_page": True,
            }], str(output_path))
            result = fitz.open(output_path)
            try:
                self.assertEqual(result.page_count, 1)
                blocks = [block for block in result[0].get_text("blocks") if "FULL PAGE" in block[4]]
                self.assertEqual(len(blocks), 1)
                self.assertGreater(blocks[0][1], 170)
            finally:
                result.close()

    def test_normal_print_keeps_selected_pages_as_separate_pages(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "source.pdf"
            source = PdfWriter()
            source.add_blank_page(width=300, height=400)
            source.add_blank_page(width=500, height=600)
            with source_path.open("wb") as stream:
                source.write(stream)

            result = pdfm_v3.PDFMergerAPI(mode="local").merge_pages_for_print([
                {"path": str(source_path), "page_index": 1, "rotation": 0},
                {"path": str(source_path), "page_index": 0, "rotation": 90},
            ])
            self.assertTrue(result["success"])
            encoded = result["data"].split(",", 1)[1]
            output_path = Path(temp_dir) / "printed.pdf"
            output_path.write_bytes(base64.b64decode(encoded))
            printed = PdfReader(str(output_path))
            self.assertEqual(len(printed.pages), 2)
            self.assertAlmostEqual(float(printed.pages[0].mediabox.width), 500)
            self.assertAlmostEqual(float(printed.pages[0].mediabox.height), 600)
            self.assertEqual(printed.pages[1].get("/Rotate"), 90)


if __name__ == "__main__":
    unittest.main()

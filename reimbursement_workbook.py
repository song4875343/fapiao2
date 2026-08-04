"""Editable workbook and PDF-print support for the double-page taxi form."""

import base64
import re
import sys
import tempfile
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, range_boundaries
FORM_NAME = "出租车报销单双页.xlsx"
MAX_ROW = 41
MAX_COLUMN = 8
DATA_BLOCKS = ((6, 17, 18, 19, 3), (27, 38, 39, 40, 24))
CELL_RE = re.compile(r"^([A-H])([1-9]|[1-3][0-9]|4[01])$")
_REPORTLAB = None


def _load_reportlab():
    """Load binary-backed PDF dependencies outside the bundled Python 3.12 path."""
    global _REPORTLAB
    if _REPORTLAB is not None:
        return _REPORTLAB
    incompatible = str(Path(__file__).resolve().parent / "baoxiao" / ".venv" / "Lib" / "site-packages")
    original_path = list(sys.path)
    try:
        sys.path[:] = [entry for entry in sys.path if str(Path(entry).resolve()) != incompatible]
        for name, module in list(sys.modules.items()):
            module_file = str(getattr(module, "__file__", "") or "")
            if (name == "PIL" or name.startswith("PIL.")) and incompatible in module_file:
                del sys.modules[name]
        from reportlab.lib.colors import Color, black
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        from reportlab.pdfgen import canvas
        _REPORTLAB = {
            "Color": Color, "black": black, "A4": A4, "pdfmetrics": pdfmetrics,
            "UnicodeCIDFont": UnicodeCIDFont, "canvas": canvas,
        }
        return _REPORTLAB
    finally:
        sys.path[:] = original_path


def _validated_path(value):
    path = Path(value).expanduser().resolve()
    if path.name != FORM_NAME or not path.is_file():
        raise ValueError("出租车报销单双页文件不存在")
    return path


def _display_value(value):
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return value


def _clear_rewrite_values(workbook):
    changed = False
    for sheet in workbook.worksheets:
        for start, end, *_ in DATA_BLOCKS:
            for row in range(start, end + 1):
                cell = sheet.cell(row, 8)
                if str(cell.value or "").strip() == "改写":
                    cell.value = None
                    changed = True
    return changed


def _normalize_form_layout(workbook):
    """Repair presentation-only template details on generated and older forms."""
    changed = False
    for sheet in workbook.worksheets:
        for coordinate in ("A2", "A23"):
            cell = sheet[coordinate]
            if cell.alignment.horizontal != "center":
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical=cell.alignment.vertical or "center",
                    wrap_text=cell.alignment.wrap_text,
                    text_rotation=cell.alignment.text_rotation,
                    shrink_to_fit=cell.alignment.shrink_to_fit,
                    indent=cell.alignment.indent,
                )
                changed = True
    return changed


def _number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _recalculate_sheet(sheet):
    populated = []
    for start, end, total_row, text_row, date_row in DATA_BLOCKS:
        rows = [row for row in range(start, end + 1) if sheet.cell(row, 6).value not in (None, "")]
        populated.append(rows)
        sheet.cell(date_row, 7).value = f"票据 {len(rows)} 张"

    all_rows = populated[0] + populated[1]
    total = round(sum(_number(sheet.cell(row, 6).value) for row in all_rows), 2)
    if populated[1]:
        sheet.cell(18, 6).value = None
        sheet.cell(19, 1).value = None
        target_total_row, target_text_row = 39, 40
    else:
        target_total_row, target_text_row = 18, 19
        sheet.cell(39, 6).value = None
        if sheet.cell(40, 1).value and "合计人民币" in str(sheet.cell(40, 1).value):
            sheet.cell(40, 1).value = None
    sheet.cell(target_total_row, 6).value = total
    sheet.cell(target_text_row, 1).value = f"合计人民币（大写）：{total:.2f}"


def _coerce_edit(cell, value):
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip()
    if cell.column in (1, 2, 6, 7) and cell.row in tuple(
        row for start, end, *_ in DATA_BLOCKS for row in range(start, end + 1)
    ):
        number = float(text)
        return int(number) if number.is_integer() and cell.column != 6 else number
    return text


def _apply_edits(workbook, edits):
    sheets = {sheet.title: sheet for sheet in workbook.worksheets}
    for edit in edits or []:
        sheet = sheets.get(str(edit.get("sheet", "")))
        address = str(edit.get("cell", "")).upper()
        if sheet is None or not CELL_RE.match(address):
            raise ValueError("报销单编辑位置无效")
        cell = sheet[address]
        if cell.coordinate in sheet.merged_cells and cell.coordinate != str(
            next(rng for rng in sheet.merged_cells.ranges if cell.coordinate in rng)
        ).split(":", 1)[0]:
            continue
        cell.value = _coerce_edit(cell, edit.get("value"))
    for sheet in workbook.worksheets:
        _recalculate_sheet(sheet)


def _merged_cell_map(sheet):
    result = {}
    for merged in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged))
        if min_row > MAX_ROW or min_col > MAX_COLUMN:
            continue
        result[(min_row, min_col)] = (
            min(max_row, MAX_ROW) - min_row + 1,
            min(max_col, MAX_COLUMN) - min_col + 1,
        )
        for row in range(min_row, min(max_row, MAX_ROW) + 1):
            for col in range(min_col, min(max_col, MAX_COLUMN) + 1):
                if (row, col) != (min_row, min_col):
                    result[(row, col)] = None
    return result


def workbook_payload(path, clear_rewrite=True):
    path = _validated_path(path)
    workbook = load_workbook(path, data_only=False)
    changed = _normalize_form_layout(workbook)
    if clear_rewrite:
        changed = _clear_rewrite_values(workbook) or changed
    if changed:
        workbook.save(path)
    sheets = []
    for sheet in workbook.worksheets:
        merged = _merged_cell_map(sheet)
        rows = []
        for row in range(1, MAX_ROW + 1):
            cells = []
            for col in range(1, MAX_COLUMN + 1):
                span = merged.get((row, col), (1, 1))
                if span is None:
                    continue
                cell = sheet.cell(row, col)
                cells.append({
                    "cell": f"{get_column_letter(col)}{row}",
                    "value": _display_value(cell.value),
                    "rowspan": span[0],
                    "colspan": span[1],
                    "bold": bool(cell.font.bold),
                    "fontSize": float(cell.font.sz or 10),
                    "align": cell.alignment.horizontal or "left",
                })
            rows.append({"height": float(sheet.row_dimensions[row].height or 15), "cells": cells})
        sheets.append({
            "name": sheet.title,
            "rows": rows,
            "columnWidths": [float(sheet.column_dimensions[get_column_letter(col)].width or 8.43) for col in range(1, MAX_COLUMN + 1)],
        })
    return {"success": True, "path": str(path), "sheets": sheets}


def save_edits(path, edits):
    path = _validated_path(path)
    workbook = load_workbook(path, data_only=False)
    _clear_rewrite_values(workbook)
    _normalize_form_layout(workbook)
    _apply_edits(workbook, edits)
    workbook.save(path)
    return workbook_payload(path, clear_rewrite=False)


def _rgb(value, default=None):
    if not value or value.type != "rgb" or not value.rgb:
        return default
    raw = value.rgb[-6:]
    try:
        return _load_reportlab()["Color"](*(int(raw[index:index + 2], 16) / 255 for index in (0, 2, 4)))
    except ValueError:
        return default


def _wrap_text(text, font_name, font_size, max_width):
    pdfmetrics = _load_reportlab()["pdfmetrics"]
    lines = []
    for paragraph in str(text).splitlines() or [""]:
        current = ""
        for char in paragraph:
            candidate = current + char
            if current and pdfmetrics.stringWidth(candidate, font_name, font_size) > max_width:
                lines.append(current)
                current = char
            else:
                current = candidate
        lines.append(current)
    return lines or [""]


def _render_sheet(pdf, sheet):
    reportlab = _load_reportlab()
    pdfmetrics = reportlab["pdfmetrics"]
    black = reportlab["black"]
    page_width, page_height = reportlab["A4"]
    col_points = [
        (sheet.column_dimensions[get_column_letter(col)].width or 8.43) * 5.25
        for col in range(1, MAX_COLUMN + 1)
    ]
    row_points = [sheet.row_dimensions[row].height or 15 for row in range(1, MAX_ROW + 1)]
    scale = min((page_width - 28) / sum(col_points), (page_height - 28) / sum(row_points))
    widths = [value * scale for value in col_points]
    heights = [value * scale for value in row_points]
    left = (page_width - sum(widths)) / 2
    top = page_height - (page_height - sum(heights)) / 2
    xs = [left]
    for width in widths:
        xs.append(xs[-1] + width)
    ys = [top]
    for height in heights:
        ys.append(ys[-1] - height)

    merged = _merged_cell_map(sheet)
    for row in range(1, MAX_ROW + 1):
        for col in range(1, MAX_COLUMN + 1):
            span = merged.get((row, col), (1, 1))
            if span is None:
                continue
            row_span, col_span = span
            cell = sheet.cell(row, col)
            x, y = xs[col - 1], ys[row - 1 + row_span]
            width = xs[col - 1 + col_span] - x
            height = ys[row - 1] - y
            fill = _rgb(cell.fill.fgColor) if cell.fill.fill_type else None
            if fill:
                pdf.setFillColor(fill)
                pdf.rect(x, y, width, height, stroke=0, fill=1)
            bordered = any(
                getattr(sheet.cell(r, c).border, side).style
                for r in range(row, row + row_span)
                for c in range(col, col + col_span)
                for side in ("left", "right", "top", "bottom")
            )
            if bordered:
                pdf.setStrokeColor(black)
                pdf.setLineWidth(max(0.35, 0.55 * scale))
                pdf.rect(x, y, width, height, stroke=1, fill=0)
            value = _display_value(cell.value)
            if value == "":
                continue
            font_size = max(5, min(18, float(cell.font.sz or 10) * scale))
            pdf.setFont("STSong-Light", font_size)
            pdf.setFillColor(_rgb(cell.font.color, black) if cell.font.color else black)
            lines = _wrap_text(value, "STSong-Light", font_size, max(4, width - 5))
            line_height = font_size * 1.15
            text_height = len(lines) * line_height
            cursor_y = y + (height + text_height) / 2 - line_height
            alignment = cell.alignment.horizontal or "left"
            for line in lines:
                line_width = pdfmetrics.stringWidth(line, "STSong-Light", font_size)
                if alignment == "center":
                    cursor_x = x + (width - line_width) / 2
                elif alignment == "right":
                    cursor_x = x + width - line_width - 3
                else:
                    cursor_x = x + 3
                pdf.drawString(cursor_x, cursor_y, line)
                cursor_y -= line_height
    pdf.showPage()


def render_pdf(path, edits=None):
    path = _validated_path(path)
    workbook = load_workbook(path, data_only=False)
    _clear_rewrite_values(workbook)
    _normalize_form_layout(workbook)
    _apply_edits(workbook, edits or [])
    workbook.save(path)
    reportlab = _load_reportlab()
    reportlab["pdfmetrics"].registerFont(reportlab["UnicodeCIDFont"]("STSong-Light"))
    stream = BytesIO()
    pdf = reportlab["canvas"].Canvas(stream, pagesize=reportlab["A4"], pageCompression=1)
    for sheet in workbook.worksheets:
        _render_sheet(pdf, sheet)
    pdf.save()
    data = stream.getvalue()
    temp_file = tempfile.NamedTemporaryFile(prefix="taxi_reimbursement_", suffix=".pdf", delete=False)
    try:
        temp_file.write(data)
        temp_path = temp_file.name
    finally:
        temp_file.close()
    return {
        "success": True,
        "path": temp_path,
        "name": "出租车报销单双页.pdf",
        "page_count": len(workbook.worksheets),
        "data": "data:application/pdf;base64," + base64.b64encode(data).decode("ascii"),
    }
